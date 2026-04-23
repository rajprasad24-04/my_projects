"""
Excel export service.
Appends risk assessment data to a company Excel file.
# STORAGE: Change EXCEL_FILE path when deploying to cloud/shared drive.
"""

from __future__ import annotations
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services.onedrive import upload_excel_to_onedrive

# STORAGE: Change this path to your company's shared Excel file location
# e.g. "C:/Company/Shared/Risk_Assessments.xlsx"
# or a OneDrive/SharePoint path
EXCEL_FILE = Path(__file__).parent.parent.parent / "data" / "risk_assessments.xlsx"

# Column headers
HEADERS = [
    "Date", "Time", "Client Name", "Time Horizon (Years)",
    "Investment Amount (₹)", "Risk Score", "Risk Category", "Risk Level",
    "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8", "Q9", "Q10"
]

# WCG Brand Colors
HEADER_BG = "1453C3"    # WCG Blue
HEADER_FG = "FFFFFF"    # White
ALT_ROW_BG = "CBDDFF"  # WCG Light Blue
BORDER_COLOR = "C5C5C5"


def _get_or_create_workbook() -> tuple[openpyxl.Workbook, any]:
    """Load existing workbook or create new one with headers."""
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Risk Assessments"
        _write_headers(ws)

    return wb, ws


def _write_headers(ws) -> None:
    """Write styled header row."""
    header_font = Font(
        name="Montserrat", bold=True,
        color=HEADER_FG, size=11
    )
    header_fill = PatternFill(
        fill_type="solid", fgColor=HEADER_BG
    )
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    col_widths = [12, 10, 20, 18, 22, 12, 14, 12] + [6] * 10
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = width

    # Set header row height
    ws.row_dimensions[1].height = 36


def _style_data_row(ws, row: int) -> None:
    """Apply alternating row colors and borders."""
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    # Alternate row color
    fill = None
    if row % 2 == 0:
        fill = PatternFill(fill_type="solid", fgColor=ALT_ROW_BG)

    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill

    # Make client name left-aligned
    ws.cell(row=row, column=3).alignment = Alignment(
        horizontal="left", vertical="center"
    )


def append_assessment(
    user_name: str,
    time_horizon: int,
    investment_amount: float,
    score: int,
    category: str,
    level: str,
    answers: dict[str, int],
    created_at: str,
) -> bool:
    """
    Append a new assessment row to the Excel file.
    Returns True if successful, False if failed.
    """
    try:
        EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
        wb, ws = _get_or_create_workbook()

        # Parse date and time
        try:
            dt = datetime.fromisoformat(created_at)
            date_str = dt.strftime("%d %b %Y")
            time_str = dt.strftime("%I:%M %p")
        except Exception:
            date_str = created_at[:10]
            time_str = ""

        # Format time horizon
        if time_horizon == 0:
            horizon_text = "< 1 year"
        elif time_horizon == 21:
            horizon_text = "> 20 years"
        else:
            horizon_text = str(time_horizon)

        # Build row data
        row_data = [
            date_str,
            time_str,
            user_name,
            horizon_text,
            round(investment_amount, 2),
            score,
            category,
            level,
            answers.get("Q1", ""),
            answers.get("Q2", ""),
            answers.get("Q3", ""),
            answers.get("Q4", ""),
            answers.get("Q5", ""),
            answers.get("Q6", ""),
            answers.get("Q7", ""),
            answers.get("Q8", ""),
            answers.get("Q9", ""),
            answers.get("Q10", ""),
        ]

        # Get next empty row
        next_row = ws.max_row + 1

        # Write data
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=next_row, column=col, value=value)

        # Style the row
        _style_data_row(ws, next_row)

        # Style score cell based on risk level
        score_colors = {
            "very-low": "1FC971",
            "low":      "86efac",
            "average":  "D1A712",
            "high":     "D67F00",
            "very-high":"CC3802",
        }
        score_color = score_colors.get(level, "1453C3")
        score_cell = ws.cell(row=next_row, column=6)
        score_cell.font = Font(bold=True, color=score_color)

        # Save file
        # Save file locally
        wb.save(EXCEL_FILE)

        # Upload to OneDrive
        upload_excel_to_onedrive(EXCEL_FILE)

        return True

    except Exception as e:
        print(f"Excel export failed: {e}")
        return False